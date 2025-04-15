from dataclasses import dataclass
from openpyxl import load_workbook


@dataclass
class ExcelHeaderData:
    column: int
    name: int


def get_excel_header_with_data():
    # Load workbook and select sheet
    workbook = load_workbook('example.xlsx')
    sheet = workbook.active  # or wb['SheetName']
    name_and_index = []
    index = 0
    for cell in sheet[1]:
        name_and_index.append(ExcelHeaderData(column=index, name=cell.value))
        index += 1
    return name_and_index


def get_excel_headers():
    # Load workbook and select sheet
    workbook = load_workbook('example.xlsx')
    sheet = workbook.active  # or wb['SheetName']
    name_and_index = []
    for cell in sheet[1]:
        name_and_index.append(cell.value)
    return name_and_index


def iterate_rows():
    name_and_data = get_excel_header_with_data()

    # Load workbook and select sheet
    workbook = load_workbook('example.xlsx')
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for data in name_and_data:
            print(f"{data.name} ({data.column}) : {row[data.column]}")


iterate_rows()
print("Headers:", get_excel_headers())
