from pathlib import Path
import math

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


class ExcelReport:

    def __init__(self, worksheet):
        self.ws = worksheet

    def setup_page(self, landscape=True):

        ws = self.ws

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = (
            ws.ORIENTATION_LANDSCAPE
            if landscape
            else ws.ORIENTATION_PORTRAIT
        )

        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.35,
            bottom=0.35,
            header=0.2,
            footer=0.2,
        )

        ws.print_options.horizontalCentered = True

    def banner(
        self,
        title,
        subtitle=None,
        report_name=None,
        logo=None,
        minimum_columns=8,
    ):

        ws = self.ws

        last_col = max(ws.max_column, minimum_columns)
        last_letter = get_column_letter(last_col)

        # -------------------------
        # Logo
        # -------------------------
        if logo and Path(logo).exists():
            img = Image(logo)
            img.width = 70
            img.height = 70
            ws.add_image(img, "A1")

        ws.column_dimensions["A"].width = 14

        # -------------------------
        # Title
        # -------------------------
        ws.merge_cells(f"B1:{last_letter}1")

        c = ws["B1"]
        c.value = title
        c.font = Font(size=18, bold=True)
        c.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        # Estimate height
        chars_per_line = max(35, last_col * 8)
        lines = math.ceil(len(title) / chars_per_line)

        ws.row_dimensions[1].height = max(28, lines * 22)

        # -------------------------
        # Subtitle
        # -------------------------
        if subtitle:

            ws.merge_cells(f"B2:{last_letter}2")

            c = ws["B2"]
            c.value = subtitle
            c.font = Font(size=13)
            c.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            ws.row_dimensions[2].height = 22

        # -------------------------
        # Report Name
        # -------------------------
        if report_name:

            ws.merge_cells(f"B3:{last_letter}3")

            c = ws["B3"]
            c.value = report_name
            c.font = Font(size=12, bold=True)
            c.alignment = Alignment(horizontal="center")

            ws.row_dimensions[3].height = 22

        ws.freeze_panes = "A5"
        ws.print_title_rows = "1:4"

    def add_table(self, headers, rows):

        ws = self.ws

        start_row = 5

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3",
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Header

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(start_row, col)

            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # Data

        for r, row in enumerate(rows, start=start_row + 1):

            for c, value in enumerate(row, start=1):

                cell = ws.cell(r, c)

                cell.value = value
                cell.border = border

                if isinstance(value, str):
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )

        self.auto_width()

    def auto_width(self):

        ws = self.ws

        for column in ws.columns:

            length = 0

            letter = column[0].column_letter

            for cell in column:

                try:
                    length = max(length, len(str(cell.value)))
                except Exception:
                    pass

            # clamp width
            width = min(max(length + 2, 10), 35)

            ws.column_dimensions[letter].width = width


# --------------------------------------------------
# Example
# --------------------------------------------------

wb = Workbook()

ws = wb.active

report = ExcelReport(ws)

report.setup_page(landscape=True)

headers = [
    "Roll",
    "Student Name",
    "Bangla",
    "English",
    "Math",
    "Science",
    "Religion",
    "ICT",
    "Total",
    "GPA",
    "Position",
]

rows = []

for i in range(1, 51):
    rows.append(
        [
            i,
            f"Student {i}",
            95,
            91,
            99,
            94,
            96,
            93,
            568,
            5.00,
            i,
        ]
    )

report.add_table(headers, rows)

report.banner(
    title="Government Bangla Fighter Higher Secondary School and College, Rangpur, Bangladesh. This is an intentionally very long title to demonstrate automatic wrapping while keeping the banner consistent across reports with different numbers of columns.",
    subtitle="Annual Examination 2026",
    report_name="Merit List",
    logo="logo.png",      # remove or change if unavailable
)

wb.save("production_report.xlsx")

print("Done")