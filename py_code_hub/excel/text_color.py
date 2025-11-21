from openpyxl import Workbook
from openpyxl.styles import Font

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Colored Text Example"

# Example data with different text colors
data = [
    ["ID", "Name", "Status"],
    [1, "Touhid Mia", "Active"],
    [2, "Fayaz Mia", "Inactive"],
    [3, "Bangla Fighter", "Pending"],
]

# Define font colors
header_font = Font(color="FF0000", bold=True)  # Red text, bold
active_font = Font(color="008000")  # Green text
inactive_font = Font(color="FF0000")  # Red text
pending_font = Font(color="FFA500")  # Orange text

# Write header
for col_num, header in enumerate(data[0], start=1):
    ws.cell(row=1, column=col_num, value=header).font = header_font

# Write data
for row_num, row in enumerate(data[1:], start=2):
    ws.cell(row=row_num, column=1, value=row[0])
    ws.cell(row=row_num, column=2, value=row[1])

    status_cell = ws.cell(row=row_num, column=3, value=row[2])
    # Set color based on status
    if row[2] == "Active":
        status_cell.font = active_font
    elif row[2] == "Inactive":
        status_cell.font = inactive_font
    elif row[2] == "Pending":
        status_cell.font = pending_font

# Save workbook
wb.save("colored_text.xlsx")
print("✅ Excel file saved: colored_text.xlsx")
